#!/usr/bin/python3

import openpyxl
import requests

BASE_URL_XIV_API_CHARACTER: str = "https://xivapi.com/character/"
GERMAN_TO_ENGLISH_CLASS_DICT: dict = {"Paladin": "Paladin",
                                      "Krieger": "Warrior",
                                      "Dunkelritter": "Dark Knight",
                                      "Revolverklinge": "Gunbreaker",
                                      "Weißmagier": "White Mage",
                                      "Gelehrter": "Scholar",
                                      "Astrologe": "Astrologian",
                                      "Mönch": "Monk",
                                      "Dragoon": "Dragoon",
                                      "Ninja": "Ninja",
                                      "Samurai": "Samurai",
                                      "Barde": "Bard",
                                      "Maschinist": "Machinist",
                                      "Tänzer": "Dancer",
                                      "Schwarzmagier": "Black Mage",
                                      "Beschwörer": "Summoner",
                                      "Rotmagier": "Red Mage",
                                      "Blaumagier": "Blue Mage (Limited Job)",
                                      "Zimmerer": "Carpenter",
                                      "Grobschmied": "Blacksmith",
                                      "Plattner": "Armorer",
                                      "Goldschmied": "Goldsmith",
                                      "Gerber": "Leatherworker",
                                      "Weber": "Weaver",
                                      "Alchemist": "Alchemist",
                                      "Mienenarbeiter": "Miner",
                                      "Gärtner": "Botanist",
                                      "Fischer": "Fisher",
                                      "Gourmet": "Culinarian"}
DEBUG_ENABLED = False


def main():
    """main method, used to process data and update the excel workbook"""

    workbook: openpyxl.Workbook = openpyxl.load_workbook("/home/amatus/Downloads/Mitgliederliste_empty.xlsx")
    worksheet = workbook.active
    class_range: tuple = generate_class_range(worksheet)
    for i in range(worksheet.min_row + 1, worksheet.max_row):
        current_row: tuple = worksheet[i]
        if not current_row[0].value and not current_row[1].value:
            break
        current_character_name: str = f"{current_row[0].value} {current_row[1].value}"
        current_character_info: dict = process_class_info(get_character_info(get_character_id(current_character_name)))
        if not current_character_info:
            print(f"Cant process data for character: {current_character_name}")
            continue
        update_character_info(current_character_info, worksheet, class_range, worksheet[worksheet.min_row], i)
    workbook.save("/home/amatus/Downloads/Mitgliederliste_updated.xlsx")


def update_character_info(current_character_info: dict, worksheet: openpyxl.workbook.workbook.Worksheet,
                          class_range: tuple, header_row: tuple, current_row: int):
    """method to update the character class information in the excel sheet"""
    for i in range(class_range[0], class_range[1]):
        mapped_class_name = GERMAN_TO_ENGLISH_CLASS_DICT.get(header_row[i].value)
        new_class_val = current_character_info.get(mapped_class_name, 0)
        if DEBUG_ENABLED:
            character_name = f"{worksheet.cell(current_row, 1).value} {worksheet.cell(current_row, 2).value}"
            current_class = header_row[i].value
            print(f"Setting value {new_class_val} for class {current_class} for character {character_name}")
        current_cell = worksheet.cell(row=current_row, column=i)
        current_cell.value = new_class_val


def process_class_info(class_info: dict):
    """method to process the class info of every player, mapping it into a dictionary for easier usage"""
    if class_info is None:
        return None

    data_to_process = class_info.get("Character", {}).get("ClassJobs", None)
    if not data_to_process:
        raise IOError
    out: dict = {entry["UnlockedState"]["Name"]: entry["Level"] for entry in data_to_process}

    if DEBUG_ENABLED:
        print("MAPPED CLASS VALUES:")
        print(out)

    return out


def generate_class_range(worksheet: openpyxl.workbook.workbook.Worksheet):
    """helper method, to create the excel ranges for the player classes"""
    header_row: tuple = worksheet[worksheet.min_row]
    end = 0
    start = 0
    start_set = False
    for col in header_row:
        if col.value is None:
            break
        if col.value in GERMAN_TO_ENGLISH_CLASS_DICT.keys() and not start_set:
            start = end
            start_set = True
        end += 1

    if DEBUG_ENABLED:
        print("CLASS ROW RANGES:")
        print(start, end)

    return start, end


def do_http_get(request_url: str):
    """helper method to do http requests"""
    resp: requests.Response = requests.get(request_url)
    if resp.ok:
        return resp.json()
    else:
        raise ConnectionError


def get_character_info(character_id: str):
    """helper method to receive character info via XIV API"""
    if not character_id:
        return None
    current_request_url: str = f"{BASE_URL_XIV_API_CHARACTER}{character_id}"
    resp_json: dict = do_http_get(current_request_url)
    return resp_json


def get_character_id(character_name: str):
    """Help method to get the ID of an character via XIV API"""
    current_request_url: str = f"{BASE_URL_XIV_API_CHARACTER}search?name={character_name}&server=Moogle"
    resp_json: dict = do_http_get(current_request_url)
    print(character_name)
    return resp_json["Results"][0]["ID"] if resp_json["Results"] else None


if __name__ == '__main__':
    main()
